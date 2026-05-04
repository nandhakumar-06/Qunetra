"""
NovaMart Flask Backend
Excel (openpyxl) as the database layer.
"""

import json, random, string
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = "data/novamart_db.xlsx"
app = Flask(__name__, static_folder="static", template_folder="templates")

# ─── helpers ────────────────────────────────────────────────────────────────

def load_wb():
    return openpyxl.load_workbook(DB_PATH)

def save_wb(wb):
    wb.save(DB_PATH)

def row_to_dict(ws, row):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {headers[i]: row[i].value for i in range(len(headers))}

def cell_border():
    s = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    return s

def gen_id(prefix, n=6):
    return prefix + "-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=n))

# ─── CORS header helper (no flask-cors needed) ───────────────────────────────

@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return response

# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory("templates", "index.html")

@app.route("/product")
def product_page():
    return send_from_directory("templates", "product.html")

@app.route("/checkout")
def checkout_page():
    return send_from_directory("templates", "checkout.html")

# GET /api/products
@app.route("/api/products", methods=["GET", "OPTIONS"])
def get_products():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    wb = load_wb()
    ws = wb["Products"]
    products = [row_to_dict(ws, ws[r]) for r in range(2, ws.max_row + 1)
                if ws.cell(r, 1).value]
    return jsonify(products)

# GET /api/products/<id>
@app.route("/api/products/<pid>", methods=["GET"])
def get_product(pid):
    wb = load_wb()
    ws = wb["Products"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == pid:
            return jsonify(row_to_dict(ws, ws[r]))
    return jsonify({"error": "Not found"}), 404

# POST /api/promo  { "code": "SAVE10" }
@app.route("/api/promo", methods=["POST", "OPTIONS"])
def validate_promo():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    code = (request.json or {}).get("code", "").strip().upper()
    wb = load_wb()
    ws = wb["PromoCodes"]
    for r in range(2, ws.max_row + 1):
        row = row_to_dict(ws, ws[r])
        if row.get("Code") == code and row.get("Active"):
            return jsonify({"valid": True, "type": row["Type"], "value": row["Value"],
                            "description": row["Description"]})
    return jsonify({"valid": False, "error": "Invalid or expired promo code"})

# POST /api/orders  { customer info + items + totals }
@app.route("/api/orders", methods=["POST", "OPTIONS"])
def place_order():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.json or {}

    # Basic validation
    required = ["name", "email", "items"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    if not data["items"]:
        return jsonify({"error": "Cart is empty"}), 400

    wb = load_wb()
    ws_p = wb["Products"]
    ws_o = wb["Orders"]
    ws_i = wb["InventoryLog"]
    ws_c = wb["PromoCodes"]

    # Recompute totals server-side (trust nothing from client)
    subtotal = 0
    order_items = []
    for item in data["items"]:
        pid, qty = item.get("id"), int(item.get("qty", 1))
        for r in range(2, ws_p.max_row + 1):
            if ws_p.cell(r, 1).value == pid:
                stock = ws_p.cell(r, 4).value or 0
                price = ws_p.cell(r, 3).value or 0
                name  = ws_p.cell(r, 2).value
                if qty > stock:
                    return jsonify({"error": f"{name} only has {stock} left in stock"}), 400
                subtotal += price * qty
                order_items.append({"id": pid, "name": name, "price": price, "qty": qty})
                break

    tax = round(subtotal * 0.18)
    shipping = 199 if subtotal > 0 else 0
    discount = 0
    promo = (data.get("promo") or "").strip().upper()

    if promo:
        for r in range(2, ws_c.max_row + 1):
            row = row_to_dict(ws_c, ws_c[r])
            if row.get("Code") == promo and row.get("Active"):
                if row["Type"] == "percent":
                    discount = round(subtotal * row["Value"] / 100)
                elif row["Type"] == "shipping":
                    shipping = 0
                elif row["Type"] == "flat":
                    discount = int(row["Value"])
                ws_c.cell(r, 5).value = (row.get("Used_Count") or 0) + 1
                break

    total = max(0, subtotal + tax + shipping - discount)

    # Write order row
    order_id = gen_id("ORD")
    auth_id  = gen_id("AUTH")
    now      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cell_font = Font(name="Arial", size=10)
    b = cell_border()
    next_row = ws_o.max_row + 1
    vals = [order_id, auth_id, now, data["name"], data["email"],
            json.dumps(order_items), subtotal, tax, shipping, discount, total, "CONFIRMED"]
    for c, v in enumerate(vals, 1):
        cell = ws_o.cell(next_row, c, value=v)
        cell.font = cell_font
        cell.border = b
        if c in (7, 8, 9, 10, 11):
            cell.number_format = "₹#,##0"

    # Deduct stock + log inventory
    log_row = ws_i.max_row + 1
    for item in order_items:
        for r in range(2, ws_p.max_row + 1):
            if ws_p.cell(r, 1).value == item["id"]:
                old_stock = ws_p.cell(r, 4).value or 0
                new_stock = max(0, old_stock - item["qty"])
                ws_p.cell(r, 4).value = new_stock
                # Inventory log
                log_vals = [now, item["id"], item["name"],
                            -item["qty"], new_stock, f"Order {order_id}"]
                for c, v in enumerate(log_vals, 1):
                    ws_i.cell(log_row, c, value=v).font = cell_font
                log_row += 1
                break

    save_wb(wb)

    return jsonify({
        "success": True,
        "order_id": order_id,
        "auth_id":  auth_id,
        "totals": {"subtotal": subtotal, "tax": tax,
                   "shipping": shipping, "discount": discount, "total": total},
        "items": order_items,
        "date": now,
    })

# GET /api/orders  — admin view
@app.route("/api/orders", methods=["GET"])
def list_orders():
    wb = load_wb()
    ws = wb["Orders"]
    orders = [row_to_dict(ws, ws[r]) for r in range(2, ws.max_row + 1)
              if ws.cell(r, 1).value]
    return jsonify(orders)

if __name__ == "__main__":
    print("🚀 NovaMart backend running → http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
