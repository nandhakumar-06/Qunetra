import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_fill = PatternFill("solid", start_color="1E3A5F")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
cell_font = Font(name="Arial", size=10)
border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

ws_p = wb.active
ws_p.title = "Products"
p_headers = ["Product_ID","Name","Price_INR","Stock","Category","Description","Image_URL","Rating","Reviews"]
for col, h in enumerate(p_headers, 1):
    c = ws_p.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

products = [
    ["P-101","Wireless Earbuds Pro",3499,15,"Audio","ANC, Bluetooth 5.3, 24h battery, IPX4 water resistant","https://images.unsplash.com/photo-1585386959984-a41552231616?q=80&w=800&auto=format&fit=crop",4.5,2341],
    ["P-102","Smartwatch Lite",2899,20,"Wearables","SpO2, heart rate, sleep tracking, 7-day battery life","https://images.unsplash.com/photo-1523275335684-37898b6baf30?q=80&w=800&auto=format&fit=crop",4.2,1876],
    ["P-103","Mechanical Keyboard 60%",4199,10,"Peripherals","Hot-swappable switches, RGB backlight, PBT double-shot keycaps","https://images.unsplash.com/photo-1618384887929-16ec33fab9ef?q=80&w=800&auto=format&fit=crop",4.7,987],
    ["P-104","USB-C GaN Charger 65W",2199,25,"Accessories","Dual port Power Delivery, foldable plug, compact travel-ready design","https://images.unsplash.com/photo-1583863788434-e58a8ee0a91e?q=80&w=800&auto=format&fit=crop",4.3,3210],
    ["P-105","1080p Webcam",2599,12,"Peripherals","Auto light correction, built-in mic, privacy shutter, plug-and-play","https://images.unsplash.com/photo-1587826080692-f439cd0b70da?q=80&w=800&auto=format&fit=crop",4.1,654],
    ["P-106","Portable SSD 1TB",5999,8,"Storage","USB 3.2 Gen2 speeds up to 1000 MB/s, ruggedized aluminum body","https://images.unsplash.com/photo-1531492746076-161ca9bcad58?q=80&w=800&auto=format&fit=crop",4.6,1123],
    ["P-107","LED Desk Lamp",1299,30,"Accessories","3 color modes, 5 brightness levels, USB-C charging port built-in","https://images.unsplash.com/photo-1507473885765-e6ed057f782c?q=80&w=800&auto=format&fit=crop",4.0,789],
    ["P-108","Noise-Cancel Headphones",6499,6,"Audio","40h battery, multipoint Bluetooth 5.2, foldable premium build","https://images.unsplash.com/photo-1505740420928-5e560c06d30e?q=80&w=800&auto=format&fit=crop",4.8,4521],
]

for r, row in enumerate(products, 2):
    for c, val in enumerate(row, 1):
        cell = ws_p.cell(row=r, column=c, value=val)
        cell.font = cell_font
        cell.border = border

ws_o = wb.create_sheet("Orders")
o_headers = ["Order_ID","Auth_ID","Date","Customer_Name","Email","Items_JSON","Subtotal","Tax","Shipping","Discount","Total","Status"]
for col, h in enumerate(o_headers, 1):
    c = ws_o.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

ws_c = wb.create_sheet("PromoCodes")
c_headers = ["Code","Type","Value","Active","Used_Count","Description"]
for col, h in enumerate(c_headers, 1):
    c = ws_c.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

promos = [
    ["SAVE10","percent",10,True,0,"10% off subtotal"],
    ["FREESHIP","shipping",0,True,0,"Free shipping on any order"],
    ["FLAT200","flat",200,True,0,"200 flat discount"],
]
for r, row in enumerate(promos, 2):
    for c, val in enumerate(row, 1):
        ws_c.cell(r, c, value=val).font = cell_font

ws_i = wb.create_sheet("InventoryLog")
i_headers = ["Timestamp","Product_ID","Product_Name","Change","New_Stock","Reason"]
for col, h in enumerate(i_headers, 1):
    c = ws_i.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

import os
os.makedirs("data", exist_ok=True)
wb.save("data/novamart_db.xlsx")
print("Excel database created")